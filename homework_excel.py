import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

workbook = openpyxl.load_workbook("""temp/homework.xlsx""")
worksheet = workbook.active
max_row = worksheet.max_row
print(max_row)
max_column = worksheet.max_column
print(max_column)


workbook_new = Workbook()
worksheet_new = workbook_new.active



var_list_1 = ["A"]

for num in range(1, 20):
    for name in var_list_1:
        row = num
        # 1 -> A, 3 -> C, 26 -> Z
        col = get_column_letter(var_list_1.index(name) + 1)
        # записываем значение в выбранную ячейку
        worksheet[f'{col}{row}'] = str(f"{name}_{num}")
        # print(f'{col}_{row}')
for num in range(1, 20):
    for name in var_list_1:
        row = num
        # 1 -> A, 3 -> C, 26 -> Z
        col = get_column_letter(var_list_1.index(name) + 1)
        # записываем значение в выбранную ячейку
        worksheet[f'B{row}'] = str(f"_{num}")
        print(f'{col}_{row}')

worksheet.delete_cols(3, max_column)

workbook.save("Temp/homework_new.xlsx")

